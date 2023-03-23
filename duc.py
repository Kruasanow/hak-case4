import openpyxl
from openpyxl import load_workbook

name = 'data4.xlsx'
def hidden_row(name):
    # Открываем Excel-файл
    wb = openpyxl.load_workbook(name)
    
    # Получаем первый лист
    ws = wb.worksheets[0]
    
    # Добавляем новую строку после 10-й строки и заполняем ее значениями
    new_row = ['Hidden', 'metadata', 'go', 'here']
    ws.append(new_row)
    
    # Скрываем новую строку
    ws.row_dimensions[ws.max_row].hidden = True
    
    # Сохраняем изменения в файле
    wb.save(name)

######################################################
# add hidden лист
######################################################
def hidden_sheet(name):
    # Открываем Excel-файл
    wb = openpyxl.load_workbook(name)
    print(wb.sheetnames)
    # Создаем новый лист
    ws = wb.create_sheet('Скрытый лист')
    
    # Скрываем лист
    ws.sheet_state = 'hidden'
    
    # Сохраняем изменения в файле
    wb.save(name)

######################################################
# поменять описание файла
######################################################
def hidden_description(name):
    # Загрузите существующий Excel файл
    file_path = name
    workbook = load_workbook(file_path)
    
    # Измените метаданные (описание) файла
    workbook.properties.title = "Новый заголовок"
    workbook.properties.author = "Новый автор"
    workbook.properties.description = "Новое описание"
    workbook.properties.subject = "Новая тема"
    workbook.properties.keywords = "ключевые слова"
    workbook.properties.lastModifiedBy = "Имя последнего редактора"
    workbook.properties.category = "Категория"
    workbook.properties.comments = "Комментарии"
    # Сохраните изменения в файле
    workbook.save(file_path)
    print("Описание Excel файла успешно обновлено.")

hidden_description(name)
hidden_sheet(name)
hidden_row(name)
