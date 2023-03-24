
# Импортируем необходимые модули
from openpyxl import load_workbook
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, Color
from openpyxl.styles.colors import WHITE
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.styles.colors import Color as ExcelColor
from openpyxl.utils import get_column_letter



# name = 'data2.xlsx'

def check_shrift(name):
    # Загружаем готовый файл Excel
    workbook = load_workbook(name)

    # Выбираем нужный лист
    worksheet = workbook.active

    # Получаем ячейку А2
    cell = worksheet['A3']

    # Получаем цвет шрифта в ячейке
    font = cell.font
    color = font.color.rgb

    # Преобразуем цвет в кортеж RGB
    rgb = str(ExcelColor(rgb=color).rgb)[2::]
    k1 = int(rgb[0:2],16)
    k2 = int(rgb[2:4],16)
    k3 = int(rgb[4:6],16)
    s = [k1,k2,k3]
    rgb = [40, 37, 44]
    word = ''
    for i in rgb:
        letter = chr(i + ord('A') - 1)
        word += letter
    return word

def hidden_sheet(name):
    # Открываем Excel-файл
    wb = openpyxl.load_workbook(name)
    return wb.sheetnames

def hidden_string(name):
    workbook = load_workbook(name)
    # Выбираем нужный лист
    worksheet = workbook.active
    # Получаем значение ячейки А2
    value = worksheet['A2'].value
    # Выводим значение на экран
    return value

def hidden_col(name):
    # Загружаем готовый файл Excel
    workbook = load_workbook(name)
    # Выбираем нужный лист
    worksheet = workbook.active
    # Получаем значения из столбца J
    column_values = []
    for cell in worksheet['J']:
        if not(cell in column_values):
            column_values.append(cell.value)
    # Выводим значения на экран
    return list(set(column_values))

def about_file(name):
    workbook = load_workbook(name)
    res = []
    res.append(f'Название файла: {name}')
    worksheet = workbook.active
    res.append(f'Имя листа: {worksheet.title}')
    res.append(f'Количество строк: {worksheet.max_row}')
    res.append(f'Количество столбцов: {worksheet.max_column}')
    res.append(str(workbook.properties))
    return res


def get_resault(name):
    resault = {
        'RESAULT with check color shrift :': check_shrift(name),
        'RESAULT hidden_sheet :': hidden_sheet(name),
        'RESAULT hidden_string :' : hidden_string(name),
        'RESAULT about file :': about_file(name),
        'RESAULT hidden_column :' : hidden_col(name),
    }

    good_keys = resault.keys()
    res = []
    for i in good_keys:
        res.append(i)
        res.append(resault[i])
    return res

# print(get_resault(name))