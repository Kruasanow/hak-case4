
######################################################
# add hidden row 0
######################################################
import openpyxl
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.styles.colors import WHITE
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.styles.colors import Color as ExcelColor
from openpyxl.utils import get_column_letter


def prepare_for_hack(name_agent1,pass_on_excel1):
    name_agent = name_agent1
    pass_on_excel = pass_on_excel1
    data_create = (str(datetime.now()).replace(':','_').replace('.','_').replace('-','_').replace(' ',''))[0:18]
    secret_message = name_agent + data_create
    return [secret_message, name_agent, pass_on_excel]


def hidden_row(name, secret_message):
    # Открываем Excel-файл
    wb = openpyxl.load_workbook(name)

    # Получаем первый лист
    ws = wb.worksheets[0]
    # ws.move_range('A2:J100', rows=1, translate=True)

    # Добавляем новую строку после 10-й строки и заполняем ее значениями
    ws.insert_rows(2)
    ws.cell(row=2, column=1).value = secret_message
    ws.row_dimensions[2].hidden = True
    font = Font(color=Color(rgb=WHITE))
    # Устанавливаем стиль шрифта в ячейке А2
    ws['A2'].font = font

    column_number = 1
    # Получаем все значения из столбца
    column_values = [cell.value for cell in ws['A'] if cell.value is not None]
    # Подсчитываем количество элементов в столбце
    count = len(column_values)
    for i in range(1,count,10):
        ws[f'J{i}'] = secret_message
        ws[f'J{i}'].font = font
    # Сохраняем изменения в файле
    wb.save(name)

######################################################
# add hidden лист
######################################################
def hidden_sheet(name, secret_message):
    # Открываем Excel-файл
    wb = openpyxl.load_workbook(name)
    print(wb.sheetnames)
    # Создаем новый лист
    ws = wb.create_sheet(secret_message)

    # Скрываем лист
    ws.sheet_state = 'hidden'
    print(wb.sheetnames)
    # Сохраняем изменения в файле
    wb.save(name)

######################################################
# поменять описание файла
######################################################
def hidden_description(name, secret_message):
    # Загрузите существующий Excel файл
    file_path = name
    workbook = load_workbook(file_path)

    # Измените метаданные (описание) файла
    workbook.properties.title = secret_message
    workbook.properties.author = secret_message
    workbook.properties.description = secret_message
    workbook.properties.subject = secret_message
    workbook.properties.keywords = secret_message
    workbook.properties.lastModifiedBy = secret_message
    workbook.properties.category = secret_message
    workbook.properties.comments = secret_message
    # Сохраните изменения в файле
    workbook.save(file_path)
    print("Описание Excel файла успешно обновлено.")

def change_color(name,name_agent):
    # Загружаем готовый файл Excel
    workbook = load_workbook(name)

    # Выбираем нужный лист
    worksheet = workbook.active


    word = name_agent
    rgb = []
    for i in word:
        number = ord(i) - ord('A') + 1
        rgb.append(number)
    k1 = rgb[0]
    k2 = rgb[1]
    k3 = rgb[2]
    rgb = (k1, k2, k3)
    rgb_fill = (255-k1, 255-k2, 255-k3)
    print(rgb_fill)
    k1 = str(hex(rgb[0]))[2::]
    k2 = str(hex(rgb[1]))[2::]
    k3 = str(hex(rgb[2]))[2::]
    color = (k1 + k2 + k3).upper()
    print(color)

    k11 = str(hex(rgb_fill[0]))[2::]
    k22 = str(hex(rgb_fill[1]))[2::]
    k33 = str(hex(rgb_fill[2]))[2::]
    color_fill = (k11 + k22 + k33).upper()
    print(color_fill, 'cplor fill')

    # Создаем объект ExcelColor с нужным цветом
    color = ExcelColor(rgb=color)
    color_fill = ExcelColor(rgb = color_fill)

    # Проходимся по всем ячейкам первого столбца и изменяем их стиль
    for row in worksheet.iter_rows(min_row=1, max_row=10000, min_col=1, max_col=1):
        cell = row[0]
        cell.font = Font(color=color)
    for row in worksheet.iter_rows(min_row=1, max_row=10000, min_col=2, max_col=2):
        cell = row[0]
        # Задаем цвет и стиль закрашивания
        fill = PatternFill(start_color=color_fill, end_color=color_fill, fill_type='solid')
        # Применяем закрашивание к ячейке
        cell.fill = fill


    # Сохраняем изменения в файл
    workbook.save(name)
    print('change color done')


def add_password(name, pass_on_excel):
    wb = openpyxl.load_workbook(name)

    wb_prot = WorkbookProtection(workbookPassword=pass_on_excel)
    wb.security = wb_prot

    # Устанавливаем пароль на каждый лист в файле (опционально)
    for sheet in wb.worksheets:
        sheet.protection.sheet = True
        sheet.protection.password = pass_on_excel
    ws = wb.worksheets[0]
    # Запрещаем копирование
    ws.protection.sheet = True
    ws.protection.allowFormatCells = False
    ws.protection.allowFormatColumns = False
    ws.protection.allowFormatRows = False
    ws.protection.allowInsertColumns = False
    ws.protection.allowInsertRows = False
    ws.protection.allowInsertHyperlinks = False
    ws.protection.allowDeleteColumns = False
    ws.protection.allowDeleteRows = False
    ws.protection.allowSort = False
    ws.protection.allowFilter = False
    ws.protection.allowUsePivotTables = False

    # Устанавливаем свойство read_only на объект Workbook
    wb.security.lockStructure = True
    wb.security.lockWindows = True
    wb.properties.data_only = True
    # Сохраняем изменения в файле
    wb.save(name)

# hidden_description(name)
# hidden_sheet(name)
# hidden_row(name)
# change_color(name,name_agent)
# add_password(name)
