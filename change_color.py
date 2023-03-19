import binascii
import openpyxl
import codecs
from openpyxl.styles import PatternFill
import datetime
from get_values import select_values

def encode(message):
    hex_message = codecs.encode(message.encode('utf-8'), 'hex').decode()
    return hex_message

def decode(message):
    string = codecs.decode(message.encode(), 'hex').decode('utf-8')
    return string

# name = select_values()[0]
# data = select_values()[1]
def change_color(name,data):
    print(name, data)
    # Зашифровать строку в 16-ричное число
    # name = "700"
    # data = '000'
    color1 = str(encode(name))
    print(color1)
    color2 = str(encode(data))
    print(color2)



    # hex_message = encode(data)
    # # Расшифровать 16-ричное число обратно в строку
    # decoded_message = binascii.unhexlify(hex_message).decode()
    # print("Расшифрованное сообщение:", decoded_message)




    # load the Excel file
    workbook = openpyxl.load_workbook('data2.xlsx')

    # select the worksheet
    worksheet = workbook['data']

    # choose the column you want to change the color of
    column = 'A'
    # choose the fill color
    fill = PatternFill(start_color=color1, end_color=color1, fill_type='solid')
    # iterate through each cell in the column and apply the fill color
    for cell in worksheet[column]:
        cell.fill = fill

    # choose the column you want to change the color of
    column = 'B'
    # choose the fill color
    fill = PatternFill(start_color=color2, end_color=color2, fill_type='solid')
    # iterate through each cell in the column and apply the fill color
    for cell in worksheet[column]:
        cell.fill = fill

    new_name_file = select_values()[0] + str(datetime.datetime.now())+'.xls'
    new_name_file = new_name_file.replace(':','_').replace(' ','_')
    a_file = 'output/' + new_name_file

    # save the changes
    workbook.save(a_file)
    print('change color works')
